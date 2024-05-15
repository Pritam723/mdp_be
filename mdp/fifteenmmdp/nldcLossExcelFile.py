import datetime
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl import Workbook
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)
import pandas as pd
import os
from openpyxl import load_workbook
from django.conf import settings


def dropIncicesFromList(listData, indices) :
    for index in sorted(indices[:-1], reverse=True):
        del listData[index]
    return listData


def createNldcLossExcel(path):

    meterFileMainFolder = os.path.join("fifteenmmdp/media/meterFile",path)

    lossConfigSheetNames = ["Drawal", "InterRegional", "Generation"]

    finalDfTranspose = []

    rgnData = pd.read_excel(meterFileMainFolder + '/Final_Output_Excel.xlsx', sheet_name = 'RGN', skiprows = 1)

    numOfDays = len(rgnData)//96
    gapColumn = [''] * (2 + numOfDays * 96)
    totalCS_Component = [0] * (numOfDays * 96)
    totalDRL_Component = [0] * (numOfDays * 96)

    # Creating the Indices that are needed to be dropped.
    dropIndices = [0]
    for i in range(1,numOfDays+1) :
        dropIndex = 99 * i - 2
        dropIndices = dropIndices + [dropIndex, dropIndex+1, dropIndex+2]

    # Appending date data 
    dateData = list(rgnData['Unnamed: 0'])
    dateData = dropIncicesFromList(dateData, dropIndices)
    dateData = ['ER','Date'] + dateData

    finalDfTranspose.append(dateData)

    # Appending Block data 
    blockData = list(rgnData['Unnamed: 1'])
    blockData = dropIncicesFromList(blockData, dropIndices)
    blocklDataToFrom = ['','Block']
    for index in range(len(blockData)) :
        if(blockData[index] == '23:45') :
            blocklDataToFrom.append(f'{blockData[index]}-24:00')
        else :
            blocklDataToFrom.append(f'{blockData[index]}-{blockData[index + 1]}')

    finalDfTranspose.append(blocklDataToFrom)


    # Filling the Drawal, Interregional and GEneration Data from RGN Sheet (rgnData) one by one. 

    for sheetName in lossConfigSheetNames :
        # lossConfig = pd.read_excel('LossCalculationConfig.xlsx', sheet_name = sheetName)

        if(os.path.exists(meterFileMainFolder+'/NPC Files/Necessary Files Local Copy/LossCalculationConfig.xlsx')) :
            print("Got the config file in local.")
            lossConfig = pd.read_excel(meterFileMainFolder+'/NPC Files/Necessary Files Local Copy/LossCalculationConfig.xlsx', sheet_name = sheetName)
        else :
            print("Got the config file in global.")

            globalNecessaryFileMainFolder = os.path.join(settings.MEDIA_ROOT + '/necessaryFiles')

            lossConfig = pd.read_excel(globalNecessaryFileMainFolder+'/LossCalculationConfig.xlsx', sheet_name = sheetName)

        
        print("Working with " + sheetName + " Data")
        for i in range(len(lossConfig)) :
            
            rowData = lossConfig.iloc[i]
            # print(rowData['NLDC Code'],rowData['Entity'],rowData['Entity Code'])
            
            nldcCode = rowData['NLDC Code']
            entity = rowData['Entity']
            entityCode = rowData['Entity Code']
            
            entityData = list(rgnData[entityCode])
            entityData = dropIncicesFromList(entityData, dropIndices)
            
            if(sheetName == "InterRegional") :
                entityData = [-item for item in entityData]
                
                
            if(sheetName == "Drawal") :

                for index, mwhValue in enumerate(entityData) :
                    # print(mwhValue)
                    # print(type(mwhValue))
                    
                    if(mwhValue < 0) :
                        totalCS_Component[index] = totalCS_Component[index] + abs(mwhValue)
                    else :
                        totalDRL_Component[index] = totalDRL_Component[index] + mwhValue
            
            if(sheetName == "InterRegional" or sheetName == "Generation") :

                for index, mwhValue in enumerate(entityData) :
                    if(mwhValue >= 0) :
                        totalCS_Component[index] = totalCS_Component[index] + mwhValue
                    else :
                        totalDRL_Component[index] = totalDRL_Component[index] + abs(mwhValue)
            
            
            entityData = [round(item,1) for item in entityData]
            entityData = [nldcCode, entity] + entityData
            finalDfTranspose.append(entityData)
            
        finalDfTranspose.append(gapColumn)

        
    # Calculating & Filling Total_CS, Total_GEN, Loss, Loss Percentage
    # print(totalCS_Component)
    # print(totalDRL_Component)

    lossData = list(map(lambda generation, drawal: round((generation - drawal),1), totalCS_Component, totalDRL_Component))
    lossDataPercentage = list(map(lambda generation, drawal: round(100*(generation - drawal)/generation,1), totalCS_Component, totalDRL_Component))

    totalCS_Component = [round(item,1) for item in totalCS_Component]
    totalCS_Component = ['', 'TOT_CS'] + totalCS_Component
    finalDfTranspose.append(totalCS_Component)


    totalDRL_Component = [round(item,1) for item in totalDRL_Component]
    totalDRL_Component = ['', 'TOT_DRL'] + totalDRL_Component
    finalDfTranspose.append(totalDRL_Component)

    lossData = ['', 'Loss'] + lossData
    finalDfTranspose.append(lossData)

    lossDataPercentage = ['','Loss Percentage'] + lossDataPercentage
    finalDfTranspose.append(lossDataPercentage)


    finalDf = [[finalDfTranspose[j][i] for j in range(len(finalDfTranspose))] for i in range(len(finalDfTranspose[0]))]

    excelData = pd.DataFrame(data = finalDf)
    # excelData.to_excel('NLDC_LOSS_FILE.xlsx', sheet_name = "NLDC Loss")

    with pd.ExcelWriter(meterFileMainFolder + '/NLDC_LOSS_FILE.xlsx', mode = 'w') as writer:
        excelData.to_excel(writer, sheet_name = "NLDC Loss")

    # colNames = []

    # for index in range(len(finalDf[0])) :
    #     if(finalDf[0][index] != finalDf[0][index] or finalDf[0][index] == '') : # isnan() check.
    #         colNames.append((finalDf[1][index], finalDf[1][index]))
    #     else :
    #         colNames.append((finalDf[0][index], finalDf[1][index]))
            
    # excelData = pd.DataFrame(data = finalDf[2:], columns = pd.MultiIndex.from_tuples(colNames))

    # excelData = excelData.set_index(excelData.columns[0])
    # excelData.to_excel('ttttttt.xlsx')

    # Making the Loss Percentage Graph

    xAxisData = dateData[1:]
    yAxisData = lossDataPercentage[1:]
    blockData = blocklDataToFrom[1:]

    book = load_workbook(meterFileMainFolder + '/NLDC_LOSS_FILE.xlsx')

    ws = book.create_sheet("Loss Percentage Graph")
        
    # write content of each row in 1st, 2nd and 3rd
    # column of the active sheet respectively.
    for index in range(len(xAxisData)):
        ws.append((xAxisData[index] + ' ' + blockData[index], yAxisData[index]))

        
    # values = Reference(ws, min_col = 1, min_row = 1,
    #                          max_col = 2, max_row = len(xAxisData))

    plotdata = Reference(ws, min_col = 2, min_row = 1, max_col = 2, max_row = len(xAxisData)+1)
    plotdates = Reference(ws, min_col=1, min_row=2,max_col=1,max_row = len(xAxisData)+1)
    
    # Chart with date axis
    c2 = LineChart()

    c2.height = 15 # default is 7.5
    c2.width = 30 # default is 

    c2.title = "Loss Percentage Graph"
    c2.style = 13
    c2.y_axis.title = "Loss Percentage"
    c2.y_axis.crossAx = 500
    c2.x_axis = DateAxis(crossAx=100)
    c2.x_axis.number_format = 'd-mmm'
    c2.x_axis.majorTimeUnit = "days"
    c2.x_axis.title = "Datetime"

    c2.add_data(plotdata, titles_from_data=True)
    c2.set_categories(plotdates)

    # Style the lines
    s0 = c2.series[0]
    s0.graphicalProperties.line.solidFill = "0000FF" # Marker filling
    # s0.smooth = True

    # s1 = c2.series[1]
    # s1.graphicalProperties.line.solidFill = "FF0000"
    # # s1.smooth = True

    ws.add_chart(c2, "E2")
    print("Job Done")

        
    book.save(meterFileMainFolder + '/NLDC_LOSS_FILE.xlsx')